from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import gspread
import datetime
import time
import urllib
from bs4 import BeautifulSoup
from urllib.request import Request
import re
import UploadFile

class ExtractProduct():

    def __init__(self):
        self.category_url = [
            {
                'url': 'https://shop.ifood.com.br/mercearia',
                'category': 'mercearia'
            },
            {
                'url': 'https://shop.ifood.com.br/bebidas',
                'category': 'bebidas'
            },
            {
                'url': 'https://shop.ifood.com.br/queijos-e-frios',
                'category': 'queijos e frios'
            },
            {
                'url': 'https://shop.ifood.com.br/acougue-e-peixaria',
                'category': 'acougue e peixaria'
            },
            {
                'url': 'https://shop.ifood.com.br/embalagens',
                'category': 'embalagens'
            },
            {
                'url': 'https://shop.ifood.com.br/mais-categorias',
                'category': 'mais categorias'
            }
        ]
        self.num_page_down = [1000, 1200, 500, 500, 1000, 500]

        self.products_num = 0

        self.excel_path = ("ProductFromShop.xlsx")  # This is file location.

        try:
            self.wb = load_workbook(self.excel_path)  # Load workbook and find active sheet.
        except:
            self.wb = Workbook()

        self.excel_sheet = self.wb.active


    def goto_mainpage(self, url):
        print("here is goto_mainpage")
        self.browser = webdriver.Chrome()
        self.browser.get(url)

        self.browser.find_element_by_class_name("button.button--large.mrg-r-16").click()
        self.browser.find_element_by_id("cep-value").send_keys("05448-000")
        self.browser.find_element_by_id("cep-button").click()
        time.sleep(3)

        return True

    def show_all_products(self, no):
        print("here is show_all_products")
        elem_body = self.browser.find_element_by_tag_name("body")

        # down the scroll of main body of page programically
        no_of_pagedowns = self.num_page_down[no]
        while no_of_pagedowns:
            elem_body.send_keys(Keys.PAGE_DOWN)
            time.sleep(0.2)
            print(no_of_pagedowns)
            no_of_pagedowns -= 1

    def extract_each_product_url_and_save(self):
        print("here is extract_each_product_url_and_save")
        self.excel_sheet.cell(1, 1).value = "URL"
        self.excel_sheet.cell(1, 2).value = "Name"
        self.excel_sheet.cell(1, 3).value = "Category"
        self.excel_sheet.cell(1, 4).value = "Price"
        self.excel_sheet.cell(1, 5).value = "Price_unit"
        self.excel_sheet.cell(1, 6).value = "Amount"
        self.excel_sheet.cell(1, 8).value = "Timestamp"
        # for i in range(7):
        #     self.excel_sheet.cell(1, i+1).fill = PatternFill(bgColor='f2ac63', fill_type = 'solid')
        self.excel_sheet.cell(2, 8).value = datetime.datetime.today().strftime("%d-%b-%Y")

        for x in range(0, 6):
            while True:
                try:
                    if (self.goto_mainpage(self.category_url[x]['url'])):
                        self.show_all_products(x)

                    # pick up information url of each product
                    products = self.browser.find_elements_by_class_name("cardMainWrap.pad-16.pad-8-mob")

                    for product in products:
                        url_each_pro = product.find_element_by_tag_name("a").get_attribute("href")
                        try:
                            name = product.find_element_by_class_name('nome').text
                            name = ' '.join(name.split())
                        except:
                            name = ''
                        try:
                            amount = ' '.join(product.find_element_by_class_name('quantidade').text.split())
                            amount = re.findall(r'[-+]?\d*\.\d+|\d+', amount)[0]
                        except:
                            amount = ''
                        try:
                            price = ' '.join(product.find_element_by_class_name('price').text.split()).replace('R$ ', '')
                        except:
                            price = ''
                        try:
                            priceperunit = ' '.join(product.find_element_by_class_name('unidades').text.split()).replace('R$ ', '')
                        except:
                            priceperunit = ''

                        # output google sheet
                        while True:
                            try:
                                self.excel_sheet.cell(self.products_num + 2, 1).value = url_each_pro
                                self.excel_sheet.cell(self.products_num + 2, 2).value = name
                                self.excel_sheet.cell(self.products_num + 2, 3).value = self.category_url[x]['category']
                                self.excel_sheet.cell(self.products_num + 2, 4).value = price
                                self.excel_sheet.cell(self.products_num + 2, 5).value = priceperunit
                                self.excel_sheet.cell(self.products_num + 2, 6).value = amount
                                print(self.products_num)
                                self.products_num += 1
                            except:
                                print("normal except")
                                continue
                            break
                except:
                    print("There is a issue from extracting each product url in the category : " + self.category_url[x]['category'])
                    continue
                break

        while True:
            try:
                self.wb.save(self.excel_path)
            except:
                print("Please check your excel file.")
                continue
            break
        self.browser.close()
        return self.products_num

if __name__ == "__main__":

    ExtractProduct = ExtractProduct()
    limit_row = ExtractProduct.extract_each_product_url_and_save()

    upload = UploadFile.UploadFile("ProductFromShop", limit_row, 8)
    upload.uploadFile()
    print("Task completed")
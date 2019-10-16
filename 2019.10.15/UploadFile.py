from __future__ import print_function
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook
import gspread

class UploadFile:
    def __init__(self, APPLICAION_NAME, LIMIT_ROW, LIMIT_COL):
        # connect to google sheet
        self.APPLICATION_NAME = APPLICAION_NAME
        self.LIMIT_ROW = LIMIT_ROW
        self.LIMIT_COL = LIMIT_COL
        scope = ['https://www.googleapis.com/auth/drive', 'https://spreadsheets.google.com/feeds']
        self.credentials = ServiceAccountCredentials.from_json_keyfile_name('SheetAndPython-8d3b50000138.json', scope)
        self.client = gspread.authorize(self.credentials)
        self.sheet = self.client.open(self.APPLICATION_NAME).sheet1

        #connect to excel file
        self.excel_path = (self.APPLICATION_NAME.replace(' ', '_') + ".xlsx")  # This is file location.
        try:
            self.wb = load_workbook(self.excel_path)  # Load workbook and find active sheet.
        except:
            self.wb = Workbook()
        self.excel_sheet = self.wb.active

    def uploadFile(self):
        data = self.copyRangeFromExcel()
        self.copyRangeToGoogle(data)

    def copyRangeFromExcel(self):
        rangeSelected = []
        # Loops through selected Rows
        for i in range(1, self.LIMIT_ROW + 1, 1):
            # Appends the row to a RowSelected list
            rowSelected = []
            for j in range(1, self.LIMIT_COL + 1, 1):
                rowSelected.append(self.excel_sheet.cell(row=i, column=j).value)
            # Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)

        return rangeSelected

    def copyRangeToGoogle(self, copiedData):
        cell_list = self.sheet.range(1, 1, self.LIMIT_ROW+1, self.LIMIT_COL)
        z = 0
        for val_i in copiedData:  # gives us a tuple of an index and value
            for val_j in val_i:
                if (val_j == None):
                    val_j = ""
                cell_list[z].value = val_j  # use the index on cell_list and the val from cell_values
                z += 1

        self.sheet.update_cells(cell_list)




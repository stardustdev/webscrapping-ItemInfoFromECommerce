from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import gspread
import datetime
import time
import urllib
from bs4 import BeautifulSoup
from urllib.request import Request
import re
import UploadFile

class ExtractProduct():

    def __init__(self):
        # set chrome webdriver
        self.main_url = "https://jtcd.com.br/produtos/"
        self.browser = webdriver.Chrome()
        self.browser.get(self.main_url)

        self.products_url = []
        self.products_num = 0

        self.page_num = 0
        self.urls = []
        self.work_finished = False

        self.excel_path = ("JTC_Bot.xlsx")  # This is file location.

        try:
            self.wb = load_workbook(self.excel_path)  # Load workbook and find active sheet.
        except:
            self.wb = Workbook()

        self.excel_sheet = self.wb.active

    def get_information_from_each_product(self):
        self.excel_sheet.cell(1, 1).value = "Name"
        self.excel_sheet.cell(1, 2).value = "Category"
        self.excel_sheet.cell(1, 3).value = "Price"
        self.excel_sheet.cell(1, 4).value = "SP"
        self.excel_sheet.cell(1, 5).value = "Packing"
        self.excel_sheet.cell(1, 7).value = "Timestamp"
        # for i in range(7):
        #     self.excel_sheet.cell(1, i+1).fill = PatternFill(bgColor='f2ac63', fill_type = 'solid')
        self.excel_sheet.cell(2, 7).value = datetime.datetime.today().strftime("%d-%b-%Y")

        self.browser.find_element_by_xpath('/html/body/div[4]/div/a').click()
        time.sleep(1)

        while True:
            try:
                all_products = self.browser.find_elements_by_class_name('add-cart-mobile')
                for product in all_products:
                    infos = product.find_elements_by_tag_name('li')

                    name = infos[1].text
                    category = infos[3].text
                    price = infos[7].text.replace('R$ ', '').replace('---', '')
                    SP = infos[9].text.replace('R$ ', '').replace('---', '')
                    packing = infos[11].find_element_by_class_name('web').text.replace('Embalagem de ', '').replace('-', '')

                    while True:
                        try:
                            self.excel_sheet.cell(self.products_num + 2, 1).value = name
                            self.excel_sheet.cell(self.products_num + 2, 2).value = category
                            self.excel_sheet.cell(self.products_num + 2, 3).value = price
                            self.excel_sheet.cell(self.products_num + 2, 4).value = SP
                            self.excel_sheet.cell(self.products_num + 2, 5).value = packing
                            print(self.products_num)
                            self.products_num += 1
                        except:
                            print("normal except")
                            continue
                        break
            except:
                continue
            break
        while True:
            try:
                self.wb.save(self.excel_path)
            except:
                print("Please check your excel file.")
                continue
            break
        self.browser.close()
        return self.products_num

if __name__ == "__main__":

    ExtractProduct = ExtractProduct()
    limit_row = ExtractProduct.get_information_from_each_product()

    upload = UploadFile.UploadFile("JTC_Bot", limit_row, 7)
    upload.uploadFile()
    print("Task completed")
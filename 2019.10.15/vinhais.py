from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import gspread
import datetime
import time
import urllib
from bs4 import BeautifulSoup
from urllib.request import Request
import re
import UploadFile

class ExtractProduct():

    def __init__(self):
        # set chrome webdriver
        self.main_url = "https://www.vinhais.com.br/Default.aspx"
        self.browser = webdriver.Chrome()
        self.browser.get(self.main_url)
        self.browser.maximize_window()
        time.sleep(2)

        self.products_url = []
        self.products_num = 0

        self.categories = []
        self.page_num = 0
        self.urls = []
        self.work_finished = False

        self.excel_path = ("Vinhais.xlsx")  # This is file location.

        try:
            self.wb = load_workbook(self.excel_path)  # Load workbook and find active sheet.
        except:
            self.wb = Workbook()

        self.excel_sheet = self.wb.active

    def get_all_products_url(self):
        # stage for getting total product counter and page counter
        self.browser.find_element_by_id('top_nav').find_element_by_id('lblAcesso').click()
        time.sleep(1)
        self.browser.find_element_by_id('ContentPlaceHolder1_txtCgc_clie').send_keys("00.461.762/0001-95")
        self.browser.find_element_by_id('ContentPlaceHolder1_txtSenha').send_keys('004617')
        self.browser.find_element_by_id('ContentPlaceHolder1_btnEntrar').click()
        time.sleep(1)

        category_body = self.browser.find_element_by_id('ContentPlaceHolder1_dtlCategorias').find_elements_by_tag_name('tr')
        for i in range(len(category_body)):
            page_number = 1
            current_page = 1
            pages_checked = False
            ele = self.browser.find_element_by_id('ContentPlaceHolder1_dtlCategorias').find_element_by_id('ContentPlaceHolder1_dtlCategorias_lnkCategoria_' + str(i))
            category = ele.text
            ele.click()
            time.sleep(1)

            while True:
                products = self.browser.find_element_by_id('ContentPlaceHolder1_dtlProdutos').find_elements_by_tag_name('td')
                for product in products:
                    try:
                        print(product.find_element_by_tag_name('a').get_attribute('href'))
                        self.products_url.append(
                            {
                                'url' : product.find_element_by_tag_name('a').get_attribute('href'),
                                'category' : category
                            }
                        )
                        self.products_num += 1
                    except:
                        print('product is not shown')
                        pass

                try:
                    pages = self.browser.find_element_by_xpath(
                        '//*[@id="main_content"]/div/div[2]/table[2]/tbody/tr[2]/td').find_elements_by_tag_name('a')

                    if (len(pages) < 13 and pages_checked == False):
                        page_number = int(pages[len(pages) - 2].text)
                        pages_checked = True
                    elif(len(pages) == 13):
                        page_number += 1

                    current_page += 1
                    if (current_page > page_number):
                        break
                    self.browser.find_element_by_id('ContentPlaceHolder1_btnNext').click()
                    time.sleep(1)
                except:
                    print('next page error')
                    pass

    def get_information_from_each_product(self):
        self.excel_sheet.cell(1, 1).value = "URL"
        self.excel_sheet.cell(1, 2).value = "Name"
        self.excel_sheet.cell(1, 3).value = "Category"
        self.excel_sheet.cell(1, 4).value = "Price"
        self.excel_sheet.cell(1, 5).value = "Package_type"
        self.excel_sheet.cell(1, 7).value = "Timestamp"
        # for i in range(7):
        #     self.excel_sheet.cell(1, i+1).fill = PatternFill(bgColor='f2ac63', fill_type = 'solid')
        self.excel_sheet.cell(2, 7).value = datetime.datetime.today().strftime("%d-%b-%Y")

        for i in range(self.products_num):
            self.browser.get(self.products_url[i]['url'])
            time.sleep(1)
            name = self.browser.find_element_by_class_name('center_content').find_element_by_id('produtinho').text
            desc = self.browser.find_element_by_class_name('center_content').find_element_by_id('descricao').text
            price = self.browser.find_element_by_class_name('center_content').find_element_by_id('ContentPlaceHolder1_lblValor').text.replace('R$ ', '')

            if ("Embalagem:" in desc):
                start = desc.find("Embalagem:")+11
                package_type = desc[start:desc.find('\n', start)]
            else:
                package_type = ''

            # output into google sheet
            while True:
                try:
                    self.excel_sheet.cell(i + 2, 1).value = self.products_url[i]['url']
                    self.excel_sheet.cell(i + 2, 2).value = name
                    self.excel_sheet.cell(i + 2, 3).value = self.products_url[i]['category']
                    self.excel_sheet.cell(i + 2, 4).value = price
                    self.excel_sheet.cell(i + 2, 5).value = package_type
                    print(str(self.products_num) + "/" + str(i))
                except:
                    print("normal except")
                    continue
                break

        while True:
            try:
                self.wb.save(self.excel_path)
            except:
                print("Please check your excel file.")
                continue
            break
        self.browser.close()
        return self.products_num+1


if __name__ == "__main__":

    ExtractProduct = ExtractProduct()
    ExtractProduct.get_all_products_url()
    limit_row = ExtractProduct.get_information_from_each_product()

    upload = UploadFile.UploadFile("Vinhais", limit_row, 7)
    upload.uploadFile()
    print("Task completed")

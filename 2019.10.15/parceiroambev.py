from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import gspread
import datetime
import time
import urllib
from bs4 import BeautifulSoup
from urllib.request import Request
import re
import UploadFile

class ExtractProduct():

    def __init__(self):
        # set chrome webdriver
        self.main_url = "https://www.parceiroambev.com.br"
        self.browser = webdriver.Chrome()
        self.browser.maximize_window()
        self.browser.get(self.main_url)

        self.products_url = []
        self.products_num = 0
        self.categories = []

        self.page_num = 0
        self.urls = []
        self.work_finished = False

        self.excel_path = ("ParceiroAmbev.xlsx")  # This is file location.

        try:
            self.wb = load_workbook(self.excel_path)  # Load workbook and find active sheet.
        except:
            self.wb = Workbook()

        self.excel_sheet = self.wb.active

    def login(self):
        self.browser.find_element_by_id("mini-login").send_keys("32197578000176")
        self.browser.find_element_by_id("mini-password").send_keys("cerva625lupe")
        self.browser.find_element_by_id('btn-login').click()
        time.sleep(2)
        self.browser.find_element_by_class_name('close.webjump-close-popup').click()

    def get_categories(self):
        self.browser.get("https://www.parceiroambev.com.br/bebidas.html")
        categories = self.browser.find_element_by_id('leftnav-tree-0').find_elements_by_tag_name('li')

        for category in categories:
            self.categories.append(category.find_element_by_tag_name('a').get_attribute('href'))

        print(self.categories)

    def get_info_of_each_product(self):
        self.excel_sheet.cell(1, 1).value = "Url"
        self.excel_sheet.cell(1, 2).value = "Name"
        self.excel_sheet.cell(1, 3).value = "Category"
        self.excel_sheet.cell(1, 4).value = "Price"
        self.excel_sheet.cell(1, 5).value = "Package_type"
        self.excel_sheet.cell(1, 6).value = "Package_unit"
        self.excel_sheet.cell(1, 8).value = "Timestamp"
        # for i in range(7):
        #     self.excel_sheet.cell(1, i+1).fill = PatternFill(bgColor='f2ac63', fill_type = 'solid')
        self.excel_sheet.cell(2, 8).value = datetime.datetime.today().strftime("%d-%b-%Y")

        for category in self.categories:
            self.browser.get(category)
            time.sleep(2)
            category_name = self.browser.find_element_by_class_name('filter-list.container-filter').find_element_by_tag_name('span').text
            print("category_name : " + category_name)
            finished = False
            while (finished == False):
                while True:
                    try:
                        products = self.browser.find_element_by_class_name('products-grid.row').find_elements_by_tag_name('li')
                        for product in products:
                            url = product.find_element_by_tag_name('a').get_attribute('href')
                            name = product.find_element_by_class_name("product-name").text
                            try:
                                description = product.find_element_by_class_name('product-description')
                                package_type = description.find_element_by_class_name('package-type').text
                                package_unit = description.find_element_by_class_name('package-qty').text
                            except:
                                package_unit = ''
                                package_type = ''

                            price = product.find_element_by_class_name('regular-price').text.replace('R$ ', '')

                            # output into google sheet
                            while True:
                                try:
                                    self.excel_sheet.cell(self.products_num + 2, 1).value = url
                                    self.excel_sheet.cell(self.products_num + 2, 2).value = name
                                    self.excel_sheet.cell(self.products_num + 2, 3).value = category_name
                                    self.excel_sheet.cell(self.products_num + 2, 4).value = price
                                    self.excel_sheet.cell(self.products_num + 2, 5).value = package_type
                                    self.excel_sheet.cell(self.products_num + 2, 6).value = package_unit
                                    print(self.products_num)
                                    self.products_num += 1
                                except:
                                    print("normal except")
                                    continue
                                break
                    except:
                        continue
                    break

                try:
                    self.browser.find_element_by_class_name('next.i-next').click()
                except:
                    finished = True
                    continue
        while True:
            try:
                self.wb.save(self.excel_path)
            except:
                print("Please check your excel file.")
                continue
            break
        self.browser.close()
        return self.products_num


if __name__ == "__main__":

    ExtractProduct = ExtractProduct()
    ExtractProduct.login()
    ExtractProduct.get_categories()
    limit_row = ExtractProduct.get_info_of_each_product()

    upload = UploadFile.UploadFile("ParceiroAmbev", limit_row, 8)
    upload.uploadFile()
    print("Task completed")
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import gspread
import datetime
import time
import urllib
from bs4 import BeautifulSoup
from urllib.request import Request
import re
import UploadFile

class ExtractProduct():

    def __init__(self):
        # set chrome webdriver
        self.main_url = "https://www.porcofeliz.com.br/tabela-de-precos"
        self.browser = webdriver.Chrome()
        self.browser.get(self.main_url)

        self.products_url = []
        self.products_num = 0

        self.page_num = 0
        self.urls = []
        self.work_finished = False

        self.excel_path = ("PORCO_FELIZ.xlsx")  # This is file location.

        try:
            self.wb = load_workbook(self.excel_path)  # Load workbook and find active sheet.
        except:
            self.wb = Workbook()

        self.excel_sheet = self.wb.active

    def login(self):
        self.browser.find_element_by_id("username").send_keys("LuigiFloki")
        self.browser.find_element_by_id("password").send_keys("LUIGIFLOKI")
        self.browser.find_element_by_xpath('//*[@id="content"]/div[3]/form/fieldset/div[4]/div/button').click()
        time.sleep(2)

    def get_information_from_each_product(self):
        self.excel_sheet.cell(1, 1).value = "Name"
        self.excel_sheet.cell(1, 2).value = "Category"
        self.excel_sheet.cell(1, 3).value = "Unit"
        self.excel_sheet.cell(1, 4).value = "Price"
        self.excel_sheet.cell(1, 5).value = "Price_unit"
        self.excel_sheet.cell(1, 7).value = "Timestamp"
        # for i in range(7):
        #     self.excel_sheet.cell(1, i+1).fill = PatternFill(bgColor='f2ac63', fill_type = 'solid')
        self.excel_sheet.cell(2, 7).value = datetime.datetime.today().strftime("%d-%b-%Y")

        while True:
            try:
                product_body = self.browser.find_element_by_xpath("//*[@id='content']/div[2]/div[2]")
                each_div = product_body.find_elements_by_tag_name("div")
                for element in each_div:
                    try:
                        classname = element.get_attribute('class')
                    except:
                        continue
                    if (classname == "categoriaTituloLista"):
                        category = element.text
                        print(category)
                    else:
                        if (classname == 'linhaProdutos'):
                            name = element.find_element_by_class_name('preco-col-01').text
                            unit = element.find_element_by_class_name('preco-col-02').get_attribute('innerHTML')
                            unit = unit[unit.find('>')+1 : unit.rfind('<')]
                            price = element.find_element_by_class_name('preco-col-04').text
                            price_unit = element.find_element_by_class_name('preco-col-03').text

                            while True:
                                try:
                                    self.excel_sheet.cell(self.products_num + 2, 1).value = name
                                    self.excel_sheet.cell(self.products_num + 2, 2).value = category
                                    self.excel_sheet.cell(self.products_num + 2, 3).value = unit
                                    self.excel_sheet.cell(self.products_num + 2, 4).value = price
                                    self.excel_sheet.cell(self.products_num + 2, 5).value = price_unit
                                    print(self.products_num)
                                    self.products_num += 1
                                except:
                                    print("normal except")
                                    continue
                                break
            except:
                continue
            break
        while True:
            try:
                self.wb.save(self.excel_path)
            except:
                print("Please check your excel file.")
                continue
            break
        self.browser.close()
        return self.products_num

if __name__ == "__main__":

    ExtractProduct = ExtractProduct()
    ExtractProduct.login()
    limit_row = ExtractProduct.get_information_from_each_product()

    upload = UploadFile.UploadFile("PORCO FELIZ", limit_row, 7)
    upload.uploadFile()
    print("Task completed")
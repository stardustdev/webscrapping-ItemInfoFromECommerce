from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import gspread
import datetime
import time
import urllib
from bs4 import BeautifulSoup
from urllib.request import Request
import re
import UploadFile
import json

class ExtractProduct():

    def __init__(self):
        # set chrome webdriver
        self.main_url = 'https://www.paodeacucar.com'
        self.category_url = "https://api.gpa.digital/pa/v3/products/categories/ecom?storeId=501&split=&showSub=true"
        self.browser = webdriver.Chrome()

        self.products_url = []
        self.products_num = 0

        self.categories = []
        self.page_num = 0
        self.urls = []
        self.work_finished = False

        self.excel_path = ("Paodeacucar.xlsx")  # This is file location.

        try:
            self.wb = load_workbook(self.excel_path)  # Load workbook and find active sheet.
        except:
            self.wb = Workbook()

        self.excel_sheet = self.wb.active

    def get_all_categories(self):
        req = Request(self.category_url)
        with urllib.request.urlopen(req) as response:
            the_page = response.read()

        temp = json.loads(the_page)
        for i in range(len(temp['content'])):
            sub_len = len(temp['content'][i]['subCategory'])
            if (sub_len == 0):
                self.categories.append(
                    {
                        'url': self.main_url + temp['content'][i]['link'],
                        'category': temp['content'][i]['name']
                    }
                )
            else:
                for j in range(sub_len):
                    self.categories.append(
                        {
                            'url':self.main_url + temp['content'][i]['subCategory'][j]['link'],
                            'category':temp['content'][i]['name']
                        }
                    )

    def convertToValidURL(self, url):
        convURL = urllib.parse.urlsplit(url)
        convURL = list(convURL)
        for i in range(len(convURL)-2):
            convURL[i+2] = urllib.parse.quote(convURL[i+2])
        convURL = urllib.parse.urlunsplit(convURL)
        return convURL

    def get_information_from_each_product(self):
        print("here is extract_each_product_url_and_save")
        self.excel_sheet.cell(1, 1).value = "URL"
        self.excel_sheet.cell(1, 2).value = "Name"
        self.excel_sheet.cell(1, 3).value = "Category"
        self.excel_sheet.cell(1, 4).value = "Price"
        self.excel_sheet.cell(1, 6).value = "Timestamp"
        # for i in range(7):
        #     self.excel_sheet.cell(1, i+1).fill = PatternFill(bgColor='f2ac63', fill_type = 'solid')
        self.excel_sheet.cell(2, 6).value = datetime.datetime.today().strftime("%d-%b-%Y")

        for category in self.categories:
            self.browser.get(category['url'])
            try:
                self.browser.maximize_window()
            except:
                pass
            time.sleep(2)

            scroll_ele = self.browser.find_element_by_tag_name("body")
            products_num = -1
            compare_products_num = 0
            while(compare_products_num != products_num):
                products_num = compare_products_num
                scroll_ele.send_keys(Keys.PAGE_DOWN)
                scroll_ele.send_keys(Keys.PAGE_DOWN)
                scroll_ele.send_keys(Keys.PAGE_DOWN)
                scroll_ele.send_keys(Keys.PAGE_DOWN)
                scroll_ele.send_keys(Keys.PAGE_DOWN)
                time.sleep(4)
                while True:
                    try:
                        compare_products_num = len(self.browser.find_element_by_tag_name('infinite-scroll').find_elements_by_class_name('thumbnail'))
                    except:
                        continue
                    break

                # self.browser.execute_script("arguments[0].scrollBy(0, 600);", scroll_ele)

            print(products_num)
            products = self.browser.find_element_by_tag_name('infinite-scroll').find_elements_by_class_name('thumbnail')
            for product in products:
                url = product.find_element_by_tag_name('a').get_attribute('href')
                name = product.find_element_by_class_name('product-description.ng-binding').text
                price1 = product.find_element_by_class_name('panel-prices.placeholder-item.ng-scope').text
                if(price1.count('R$') == 2):
                    price = price1[price1.rfind('R$')+2 : price1.find('\n', price1.rfind('R$'))+1]
                    if (price == ''):
                        price = price1[price1.rfind('R$') + 3: len(price1)]
                elif (price1.count('R$') == 1):
                    try:
                        price = product.find_element_by_class_name('panel-prices.placeholder-item.ng-scope').find_element_by_class_name('normal-price.ng-binding.ng-scope').text.replace('R$ ', '')
                    except:
                        price = ''
                else:
                    price = ''

                print(name + "/" + price)


                # output google sheet
                while True:
                    try:
                        self.excel_sheet.cell(self.products_num + 2, 1).value = url
                        self.excel_sheet.cell(self.products_num + 2, 2).value = name
                        self.excel_sheet.cell(self.products_num + 2, 3).value = category['category']
                        self.excel_sheet.cell(self.products_num + 2, 4).value = price
                        print(self.products_num)
                        self.products_num += 1
                    except:
                        print("normal except")
                        continue
                    break
            print("--------------------------------------------")
        while True:
            try:
                self.wb.save(self.excel_path)
            except:
                print("Please check your excel file.")
                continue
            break
        self.browser.close()
        return self.products_num

if __name__ == "__main__":

    ExtractProduct = ExtractProduct()
    ExtractProduct.get_all_categories()
    limit_row = ExtractProduct.get_information_from_each_product()

    upload = UploadFile.UploadFile("Paodeacucar", limit_row, 6)
    upload.uploadFile()
    print("Task completed")

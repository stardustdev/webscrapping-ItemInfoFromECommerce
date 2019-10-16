import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import gspread
import datetime
import time
import urllib
from bs4 import BeautifulSoup
from urllib.request import Request
import re
import UploadFile

class ExtractProduct():

    def __init__(self):
        # set chrome webdriver
        self.browser = webdriver.Chrome()

        self.products_url = []
        self.products_num = 0

        self.categories = [
            {
                'url':'https://www.casamarepescados.com.br/frutos-do-mar.html?limit=36&mode=grid',
                'name':'FRUTOS DO MAR'
            },
            {
                'url':'https://www.casamarepescados.com.br/peixes.html?limit=36&mode=grid',
                'name':'PEIXES'
            }
        ]
        self.page_num = 0
        self.urls = []
        self.work_finished = False

        self.excel_path = ("Casamare.xlsx")  # This is file location.

        try:
            self.wb = load_workbook(self.excel_path)  # Load workbook and find active sheet.
        except:
            self.wb = Workbook()

        self.excel_sheet = self.wb.active

    def get_information_from_each_product(self):
        self.excel_sheet.cell(1, 1).value = "URL"
        self.excel_sheet.cell(1, 2).value = "Name"
        self.excel_sheet.cell(1, 3).value = "Category"
        self.excel_sheet.cell(1, 4).value = "Price_Type"
        self.excel_sheet.cell(1, 5).value = "Price"
        self.excel_sheet.cell(1, 7).value = "Timestamp"
        # for i in range(7):
        #     self.excel_sheet.cell(1, i+1).fill = PatternFill(bgColor='f2ac63', fill_type = 'solid')
        self.excel_sheet.cell(2, 7).value = datetime.datetime.today().strftime("%d-%b-%Y")

        for category in self.categories:
            self.browser.get(category['url'])
            time.sleep(1)
            while True:
                try:
                    self.browser.maximize_window()
                except:
                    continue
                break
            time.sleep(2)

            try:
                text = self.browser.find_element_by_class_name('category-products').find_element_by_class_name('amount').text
                products_num = int(text[text.rfind('de')+3 : text.find('no')-1])
                page_num = (products_num+35) // 36
            except:
                page_num = 1
                pass

            print(str(page_num))
            for i in range(page_num):
                products = self.browser.find_element_by_class_name('category-products').find_elements_by_class_name('col-xs-12.col-sm-2')
                for product in products:
                    url = product.find_element_by_tag_name('a').get_attribute('href')
                    name = product.find_element_by_class_name('product-name').text
                    price = product.find_element_by_class_name('price-box').text
                    if (price.find('R') != -1):
                        money_type = 'R$'
                    else:
                        money_type = '$'
                    price = re.findall(r'[-+]?\d*\.\d+|\d+', price)[0]
                    print(name)

                    # output google sheet
                    while True:
                        try:
                            self.excel_sheet.cell(self.products_num + 2, 1).value = url
                            self.excel_sheet.cell(self.products_num + 2, 2).value = name
                            self.excel_sheet.cell(self.products_num + 2, 3).value = category['name']
                            self.excel_sheet.cell(self.products_num + 2, 4).value = money_type
                            self.excel_sheet.cell(self.products_num + 2, 5).value = price
                            print(self.products_num)
                            self.products_num += 1
                        except:
                            print("normal except")
                            continue
                        break

                if (page_num > 1 and i < page_num-1):
                    scroll_ele = self.browser.find_element_by_tag_name('body')
                    scroll_ele.send_keys(Keys.PAGE_DOWN)
                    scroll_ele.send_keys(Keys.PAGE_DOWN)
                    scroll_ele.send_keys(Keys.PAGE_DOWN)
                    time.sleep(0.3)
                    self.browser.find_element_by_class_name('toolbar-bottom').find_element_by_class_name('next.i-next.fa.fa-caret-right').click()
                    time.sleep(1)

        while True:
            try:
                self.wb.save(self.excel_path)
            except:
                print("Please check your excel file.")
                continue
            break
        self.browser.close()
        return self.products_num


if __name__ == "__main__":

    ExtractProduct = ExtractProduct()
    limit_row = ExtractProduct.get_information_from_each_product()

    upload = UploadFile.UploadFile("Casamare", limit_row, 7)
    upload.uploadFile()
    print("Task completed")

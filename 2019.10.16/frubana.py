from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import gspread
import datetime
import time
import urllib
from bs4 import BeautifulSoup
from urllib.request import Request
import re
import UploadFile

class ExtractProduct():

    def __init__(self):
        # set chrome webdriver
        self.main_url = "https://www.frubana.com/br/home"
        self.browser = webdriver.Chrome()
        self.browser.get(self.main_url)
        self.browser.maximize_window()
        time.sleep(2)

        self.products_url = []
        self.products_num = 0

        self.categories = ['Frutas', 'Verduras', 'Raizes', 'Folhas']
        self.page_num = 0
        self.urls = []
        self.work_finished = False

        self.excel_path = ("Frubana.xlsx")  # This is file location.

        try:
            self.wb = load_workbook(self.excel_path)  # Load workbook and find active sheet.
        except:
            self.wb = Workbook()

        self.excel_sheet = self.wb.active

    def get_information_from_each_product(self):
        self.browser.find_element_by_class_name('ant-btn.region-button').click()
        time.sleep(5)

        self.excel_sheet.cell(1, 1).value = "Name"
        self.excel_sheet.cell(1, 2).value = "Category"
        self.excel_sheet.cell(1, 3).value = "Price"
        self.excel_sheet.cell(1, 4).value = "Unit"
        self.excel_sheet.cell(1, 5).value = "Amount"
        self.excel_sheet.cell(1, 7).value = "Timestamp"
        # for i in range(7):
        #     self.excel_sheet.cell(1, i+1).fill = PatternFill(bgColor='f2ac63', fill_type = 'solid')
        self.excel_sheet.cell(2, 7).value = datetime.datetime.today().strftime("%d-%b-%Y")

        for index, category in enumerate(self.categories):
            while True:
                try:
                    self.browser.find_element_by_class_name('catalogue-categories-container').find_elements_by_class_name('slick-slide.slick-active')[index+1].click()
                    time.sleep(1)
                except:
                    continue
                break
            time.sleep(2)
            products = self.browser.find_elements_by_class_name('sc-cJSrbW.dqRttR')
            for product in products:
                while True:
                    try:
                        product.click()
                        time.sleep(1)
                    except:
                        continue
                    break
                product_body = self.browser.find_element_by_class_name('ant-modal-body')
                name = product_body.find_element_by_tag_name('h3').text
                unit = product_body.find_element_by_tag_name('h4').text.replace('Unidade de vendas: ', '')
                try:
                    amount = product_body.find_element_by_class_name('calc').text.replace('Aprox.: ', '')
                except:
                    amount = ''
                price = product_body.find_element_by_class_name('sc-jlyJG.dKaQpy').text
                price = re.findall(r'[-+]?\d*,\d+|\d+', price)[0]
                # output into google sheet
                while True:
                    try:
                        self.excel_sheet.cell(self.products_num + 2, 1).value = name
                        self.excel_sheet.cell(self.products_num + 2, 2).value = category
                        self.excel_sheet.cell(self.products_num + 2, 3).value = price
                        self.excel_sheet.cell(self.products_num + 2, 4).value = unit
                        self.excel_sheet.cell(self.products_num + 2, 5).value = amount
                        print(self.products_num)
                        self.products_num += 1
                        self.browser.find_element_by_class_name('ant-modal-close').click()
                        time.sleep(1)
                    except:
                        print("normal except")
                        continue
                    break

        while True:
            try:
                self.wb.save(self.excel_path)
            except:
                print("Please check your excel file.")
                continue
            break
        self.browser.close()
        return self.products_num

if __name__ == "__main__":

    ExtractProduct = ExtractProduct()
    limit_rows = ExtractProduct.get_information_from_each_product()

    upload = UploadFile.UploadFile("Frubana", limit_rows, 7)
    upload.uploadFile()
    print("Task completed")

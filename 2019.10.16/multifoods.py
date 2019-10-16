from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import gspread
import datetime
import time
import urllib
from bs4 import BeautifulSoup
from urllib.request import Request
import re
import UploadFile

class ExtractProduct():

    def __init__(self):
        # set chrome webdriver
        self.main_url = "https://www.multifoods.com.br/produtos"
        self.browser = webdriver.Chrome()
        self.browser.get(self.main_url)
        self.browser.maximize_window()

        # parameters
        self.products_url = []
        self.products_num = 0

        self.page_num = 0
        self.urls = []
        self.work_finished = False

        self.excel_path = ("MultifoodsBOT.xlsx")  # This is file location.

        try:
            self.wb = load_workbook(self.excel_path)  # Load workbook and find active sheet.
        except:
            self.wb = Workbook()

        self.excel_sheet = self.wb.active

    def get_total_num_of_products(self):

        # stage for login
        time.sleep(5)

        self.browser.find_element_by_partial_link_text("Faça o seu login").click()
        time.sleep(2)
        while True:
            try:
                self.browser.find_element_by_class_name("simple-field").send_keys("floki.tech.contato@gmail.com")
                self.browser.find_element_by_name("senha").send_keys("flokitech2019")
                self.browser.find_element_by_xpath("/html/body/div/div[1]/div[2]/div/div[1]/div/div[1]/div/form/div[2]/input").click()

                time.sleep(4)
            except:
                continue
            break

        while True:
            try:
                # stage for getting total product counter and page counter
                total_number = self.browser.find_element_by_class_name("description").text
                self.products_num = int(re.findall(r'[-+]?\d*\.\d+|\d+', total_number)[2])
                self.page_num = int((self.products_num + 59) / 60)
                print(
                    "product number : " + str(self.products_num) + " and " + "page number : " + str(self.page_num))
            except:
                continue
            break

    def get_url_of_each_product(self):
        for i in range(0, self.page_num):
            while True:
                try:
                    elements = self.browser.find_elements_by_class_name("col-md-3.col-sm-4.shop-grid-item")
                    for ele in elements:
                        # get url of each product
                        category = ele.find_element_by_class_name("tag.categoria")
                        url = category.get_attribute("href")
                        category = category.text
                        self.products_url.append({
                            'url':url,
                            'category':category
                        })
                except:
                    print("Here is a problem.")
                    continue
                break

            # click pagination button to go to next page
            if (i != self.page_num - 1):
                while True:
                    try:
                        self.browser.find_element_by_css_selector(".pages-box > a[data-index*='" + str(i + 2) + "']").click()
                        time.sleep(3)
                    except:
                        continue
                    break
        print("test:" + str(self.products_num) + "/" + str(len(self.products_url)))

    def get_information_from_each_product(self):
        print("here is extract_info_from_each_product")
        self.excel_sheet.cell(1, 1).value = "URl"
        self.excel_sheet.cell(1, 2).value = "Name"
        self.excel_sheet.cell(1, 3).value = "Category"
        self.excel_sheet.cell(1, 4).value = "Price"
        self.excel_sheet.cell(1, 5).value = "PricePerUnit"
        self.excel_sheet.cell(1, 6).value = "Unit"
        self.excel_sheet.cell(1, 7).value = "Amount"
        self.excel_sheet.cell(1, 8).value = "Marca"
        self.excel_sheet.cell(1, 9).value = "Code"
        self.excel_sheet.cell(1, 10).value = "Barcode"
        self.excel_sheet.cell(1, 12).value = "Timestamp"
        # for i in range(7):
        #     self.excel_sheet.cell(1, i+1).fill = PatternFill(bgColor='f2ac63', fill_type = 'solid')
        self.excel_sheet.cell(2, 12).value = datetime.datetime.today().strftime("%d-%b-%Y")

        url_index = 0

        while url_index < self.products_num:
            while True:
                try:
                    try:
                        url = self.products_url[url_index]['url']
                    except:
                        self.work_finished = True
                        break

                    if (url == None) or (url == ""):  # If url doesn't exist, iteration is finished
                        self.work_finished = True
                        break

                    print(url)

                    self.browser.get(url)
                    self.browser.maximize_window()
                    time.sleep(2.5)

                    product_body = self.browser.find_element_by_class_name('product-detail-box')

                    try:
                        name = product_body.find_element_by_class_name("product-title").text
                    except:
                        name = ''

                    try:
                        code = self.browser.find_element_by_xpath("//*[contains(text(), 'Código: ')]").text.replace("Código: ", "")
                    except:
                        code = ""

                    try:
                        barcode = self.browser.find_element_by_xpath("//*[contains(text(), 'Código de Barras: ')]").text.replace("Código de Barras: ", "")
                    except:
                        barcode = ""

                    try:
                        marca = self.browser.find_element_by_xpath("//*[contains(text(), 'Marca: ')]").text.replace("Marca: ", "")
                    except:
                        marca = ""

                    try:
                        unit = self.browser.find_element_by_xpath("//*[contains(text(), 'Unidade: ')]").text.replace("Unidade: ", "")
                    except:
                        unit = ""

                    try:
                        amount = self.browser.find_element_by_xpath("//*[contains(text(), 'Peso Aproximado: ')]").text.replace("Peso Aproximado: ", "")
                    except:
                        amount = ""

                    try:
                        priceperunit = self.browser.find_element_by_xpath("//*[contains(text(), 'Preço por')]").text
                        priceperunit = re.findall(r'[-+]?\d*,\d+|\d+', priceperunit)[0]
                    except:
                        priceperunit = ""

                    try:
                        price = product_body.find_element_by_class_name("current").text.replace('R$ ', '')
                    except:
                        price = ""

                    # output into google sheet
                    while True:
                        try:
                            self.excel_sheet.cell(url_index + 2, 1).value = self.products_url[url_index]['url']
                            self.excel_sheet.cell(url_index + 2, 2).value = name
                            self.excel_sheet.cell(url_index + 2, 3).value = self.products_url[url_index]['category']
                            self.excel_sheet.cell(url_index + 2, 4).value = price
                            self.excel_sheet.cell(url_index + 2, 5).value = priceperunit
                            self.excel_sheet.cell(url_index + 2, 6).value = unit
                            self.excel_sheet.cell(url_index + 2, 7).value = amount
                            self.excel_sheet.cell(url_index + 2, 8).value = marca
                            self.excel_sheet.cell(url_index + 2, 9).value = code
                            self.excel_sheet.cell(url_index + 2, 10).value = barcode
                            print(str(self.products_num) + "/" + str(url_index + 1))
                            url_index += 1
                        except:
                            print("normal except")
                            continue
                        break

                except:
                    print("normal except_2")
                    continue
                break
        while True:
            try:
                self.wb.save(self.excel_path)
            except:
                print("Please check your excel file.")
                continue
            break
        self.browser.close()
        return url_index


if __name__ == "__main__":

    ExtractProduct = ExtractProduct()
    ExtractProduct.get_total_num_of_products()
    ExtractProduct.get_url_of_each_product()
    limit_row = ExtractProduct.get_information_from_each_product()

    upload = UploadFile.UploadFile("MultifoodsBOT", limit_row, 12)
    upload.uploadFile()
    print("Task completed")
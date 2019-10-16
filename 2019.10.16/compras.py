from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import gspread
import datetime
import time
import urllib
from bs4 import BeautifulSoup
from urllib.request import Request
import re
import UploadFile

class ExtractProduct():

    def __init__(self):
        # set chrome webdriver
        self.main_url = "https://compras.estoqueonlinealimentos.com.br/IECommerce/produtos"
        self.browser = webdriver.Chrome()
        self.browser.get(self.main_url)
        self.browser.maximize_window()
        self.browser.refresh()
        time.sleep(2)

        # parameters
        self.categories = ['ATOMATADOS', 'BEBIDAS', 'BOMBONIERE',
                         'CEREAIS', 'CHAS', 'CONDIMENTOS',
                         'CONSERVAS', 'DIET', 'DOCES',
                         'ENLATADOS', 'ESSENCIAS', 'FARINACEOS',
                         'FRUTAS SECAS / CALDA', 'GRAOS SECOS', 'LEITE E DERIVADOS',
                         'MAIONESES', 'MARGARINAS', 'MASSAS',
                         'MATINAIS', 'MOLHOS', 'OLEOS E AZEITES',
                         'PANIFICADORA E CONF', 'SACHET', 'SEMENTES',
                         'SNACKS', 'SOBREMESAS']

        self.products_url = []
        self.products_num = 0

        self.page_num = 0
        self.urls = []
        self.work_finished = False

        self.excel_path = ("EstoqueOnlineBOT.xlsx")  # This is file location.

        try:
            self.wb = load_workbook(self.excel_path)  # Load workbook and find active sheet.
        except:
            self.wb = Workbook()

        self.excel_sheet = self.wb.active

    def get_information_from_each_product(self):
        self.excel_sheet.cell(1, 1).value = "Name"
        self.excel_sheet.cell(1, 2).value = "Category"
        self.excel_sheet.cell(1, 3).value = "Price"
        self.excel_sheet.cell(1, 4).value = "Package"
        self.excel_sheet.cell(1, 5).value = "Code"
        self.excel_sheet.cell(1, 5).value = "EAN"
        self.excel_sheet.cell(1, 8).value = "Timestamp"
        # for i in range(7):
        #     self.excel_sheet.cell(1, i+1).fill = PatternFill(bgColor='f2ac63', fill_type = 'solid')
        self.excel_sheet.cell(2, 8).value = datetime.datetime.today().strftime("%d-%b-%Y")

        for category in self.categories:
            self.browser.find_element_by_link_text(category).click()
            time.sleep(1)

            scroll_ele = self.browser.find_element_by_xpath(
                "//div[@class='ui-datascroller-content ui-widget-content ui-corner-bottom']")
            # # time.sleep(1)
            no_of_pagedowns = 40

            while no_of_pagedowns:
                self.browser.execute_script("arguments[0].scrollBy(0, 600);", scroll_ele)
                time.sleep(0.2)
                no_of_pagedowns -= 1

            all_products = self.browser.find_elements_by_class_name("w3-col.l2.m6.s12.card")
            for i in range(len(all_products)):
                # # time.sleep(1)

                while True:
                    try:
                        scroll_ele = self.browser.find_element_by_xpath(
                            "//div[@class='ui-datascroller-content ui-widget-content ui-corner-bottom']")
                        no_of_pagedowns = (i//6)*10
                        while no_of_pagedowns:
                            self.browser.execute_script("arguments[0].scrollBy(0, 600);", scroll_ele)
                            time.sleep(0.2)
                            no_of_pagedowns -= 1

                        product = self.browser.find_elements_by_class_name("w3-col.l2.m6.s12.card")[i]
                        product.find_element_by_tag_name('a').click()
                        time.sleep(1)
                        infos = self.browser.find_element_by_class_name("ui-tabs-panel.ui-widget-content.ui-corner-bottom").text.split("\n")
                        print(infos)
                    except:
                        print("Here is a problem.")
                        try:
                            check = self.browser.find_element_by_id('frm_Description')
                            print("back")
                            self.browser.back()
                            time.sleep(1)
                            continue
                        except:
                            continue
                    break

                # output google sheet
                while True:
                    try:
                        self.excel_sheet.cell(self.products_num + 2, 1).value = infos[0]
                        self.excel_sheet.cell(self.products_num + 2, 2).value = category
                        self.excel_sheet.cell(self.products_num + 2, 3).value = infos[4].replace('Valor Unitário: R$ ', '')
                        self.excel_sheet.cell(self.products_num + 2, 4).value = infos[3].replace('Embalagem de Vendas: ', '')
                        self.excel_sheet.cell(self.products_num + 2, 5).value = infos[1].replace('Código do Produto: ', '')
                        self.excel_sheet.cell(self.products_num + 2, 5).value = infos[2].replace('EAN: ', '')
                        print(self.products_num)
                        self.products_num += 1
                    except:
                        print("normal except")
                        continue
                    break
                while True:
                    try:
                        # self.browser.execute_script("window.history.go(-1)")
                        self.browser.back()
                        time.sleep(2)
                    except:
                        print("back error")
                        continue
                    break

        while True:
            try:
                self.wb.save(self.excel_path)
            except:
                print("Please check your excel file.")
                continue
            break
        self.browser.close()
        return self.products_num

if __name__ == "__main__":

    ExtractProduct = ExtractProduct()
    limit_row = ExtractProduct.get_information_from_each_product()

    upload = UploadFile.UploadFile("EstoqueOnlineBOT", limit_row, 7)
    upload.uploadFile()
    print("Task completed")

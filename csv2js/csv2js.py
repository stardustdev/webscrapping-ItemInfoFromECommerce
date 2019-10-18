from bs4 import BeautifulSoup
from urllib.request import urlopen
from openpyxl import load_workbook
import re

filepath = ("1.xlsx")        # This is file location.

wb = load_workbook(filepath)    # Load workbook and find active sheet.
sheet = wb.active

entity = "export class CustomersTable {\n\tpublic static customers: any = [\n"
i = 1
while i > 0:
    entity = entity + "\t\t{\n"
    entity = entity + "\t\t\tid: " + i.__str__() + ",\n"
    entity = entity + "\t\t\tfirstName: '" + sheet.cell(i,1).value.__str__() + "',\n"
    entity = entity + "\t\t\tlastName: '" + sheet.cell(i, 2).value + "',\n"
    entity = entity + "\t\t\temail: '" + sheet.cell(i, 3).value + "',\n"
    entity = entity + "\t\t\tuserName: 'sgabotti0',\n"
    entity = entity + "\t\t\tgender: '" + sheet.cell(i, 4).value.__str__() + "',\n"
    entity = entity + "\t\t\tstatus: '" + sheet.cell(i, 5).value.strftime('%m/%d/%Y') + "',\n"
    entity = entity + "\t\t\tdateOfBbirth: '" + sheet.cell(i, 6).value.strftime('%H:%M:%S') + "',\n"
    entity = entity + "\t\t\tipAddress: 'sgabotti0',\n"
    entity = entity + "\t\t\ttype: '" + sheet.cell(i, 7).value.strftime('%H:%M:%S') + "',\n"
    entity = entity + "\t\t\t_userid: 'sgabotti0',\n"
    entity = entity + "\t\t\t_createdDate: 'sgabotti0',\n"
    entity = entity + "\t\t\t_updatedDate: 'sgabotti0'\n"
    entity = entity + "\t\t}"

    if (i == 100):
        break
    else:
        entity = entity + ",\n"

    i = i + 1

entity = entity + "\t];\n}"

my_data_file = open('data.txt', 'w')
my_data_file.write(entity)
